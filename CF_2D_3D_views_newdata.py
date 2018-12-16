
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import os
import CF_dtw 
from CF_dtw import path_cost
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.patches import FancyArrowPatch
from mpl_toolkits.mplot3d import proj3d
from scipy.fftpack import fft


def read_cs_abl_2d_data():
    wb = load_workbook('IM53.xlsx')
    sheet_1 = wb.get_sheet_by_name('sheet_1')
    wb_cs = load_workbook('IM53-cs.xlsx')
    sheet_1_cs = wb_cs.get_sheet_by_name('sheet_1')
 
    x = np.zeros(sheet_1.max_row-1) #X_BP   
    y = np.zeros(sheet_1.max_row-1) #Y_BP
    x_cs = np.zeros(sheet_1_cs.max_row-1) #X_BP   
    y_cs = np.zeros(sheet_1_cs.max_row-1) #Y_BP
 
    for i in range(1,sheet_1.max_row):
        x[i-1]=sheet_1.cell(row=i+1, column=2).value
        y[i-1]=sheet_1.cell(row=i+1, column=3).value
    for i in range(1,sheet_1_cs.max_row):
        x_cs[i-1]=sheet_1_cs.cell(row=i+1, column=2).value
        y_cs[i-1]=sheet_1_cs.cell(row=i+1, column=3).value
    return x,y,x_cs,y_cs


def read_2d_data():
    wb = load_workbook('IM53.xlsx')
    sheet_1 = wb.get_sheet_by_name('sheet_1')
 
    x = np.zeros(sheet_1.max_row-1) #X_BP   
    y = np.zeros(sheet_1.max_row-1) #Y_BP
 
    for i in range(1,sheet_1.max_row):
        x[i-1]=sheet_1.cell(row=i+1, column=2).value
        y[i-1]=sheet_1.cell(row=i+1, column=3).value
    return x,y
    
def read_3d_data():
    f = open("/Users/xuyong/Desktop/bp4-5.txt", "r")
    lines = f.readlines()[1:]
    f.close()
    x = []
    y = []
    z = []
    for line in lines:
        parts = line.split()
        x.append(float(parts[0]))
        y.append(float(parts[1]))
        z.append(float(parts[2]))
    return x,y,z

def max_maxID_segment(segment):
    tmp = -10000
    for i in range(0,len(segment)):
        if(segment[i]>tmp):
            max_seg = segment[i]
            maxID_seg = i
            tmp = segment[i]
    return max_seg, maxID_seg
    
def x_base_graph(x):
    base =  np.arange(0,len(x),1)
    fig = plt.figure()
    plt.plot(base,x,'ro-')
    plt.grid(True)
    plt.xlabel('Seq.')
    plt.ylabel('Value')
    plt.show()
    
def max_maxID(x):
    seq_peak = []
    seq_peak_id = []
    segment_element_id = 0
    up_down = -1
    segment = []
    for i in range(0, len(x)):
        if(x[i]>np.mean(x)):
            segment.append(x[i])  
            segment_element_id = i
            up_down = 1 
        else:
            if(up_down == 1):
                #seq.append(np.max(temp))
                len_segment = len(segment)
                max_segment,maxID_segment = max_maxID_segment(segment)
                seq_peak.append(max_segment)
                seq_peak_id.append(segment_element_id - len_segment + maxID_segment + 2)
                segment = []
                up_down = -1
    return seq_peak, seq_peak_id
                
def min_minID_segment(segment):
    tmp = 10000
    for i in range(0,len(segment)):
        if(segment[i]<tmp):
            min_seg = segment[i]
            minID_seg = i
            tmp = segment[i]
    return min_seg, minID_seg

def min_minID(x):
    seq_peak = []
    seq_peak_id = []
    segment_element_id = 0
    up_down = 1 #-1 down, +1 up
    segment = []
    for i in range(0, len(x)):
        if(x[i]<np.mean(x)):
            segment.append(x[i])  
            segment_element_id = i
            up_down = -1 
        else:
            if(up_down == -1):
                #seq.append(np.max(temp))
                len_segment = len(segment)
                min_segment,minID_segment = min_minID_segment(segment)
                seq_peak.append(min_segment)
                seq_peak_id.append(segment_element_id - len_segment + minID_segment + 2)
                segment = []
                up_down = +1
    return seq_peak, seq_peak_id
                      
def oneD_view_static(x,y):
    base = np.arange(1,len(x)+1,1)  
    fig = plt.figure()
    plt.subplot(211)
    plt.plot(base,x,'o-')
    plt.xlabel('base')
    plt.ylabel('X')
    plt.subplot(212)
    plt.plot(base,y,'*-')
    plt.xlabel('base')
    plt.ylabel('Y')
    plt.grid(True)
    plt.show()
    
def oneD_view_static_cs(x,y,x_cs,y_cs):
    x = x-np.mean(x)
    y = y-np.mean(y)
    x_cs = x_cs-np.mean(x_cs)
    y_cs = y_cs-np.mean(y_cs)
    base = np.arange(1,len(x)+1,1)  
    fig = plt.figure()
    plt.subplot(211)
    plt.plot(base,x,'o-')
    plt.plot(base,x_cs,'*-')
    plt.xlabel('base')
    plt.ylabel('X')
    plt.subplot(212)
    plt.plot(base,y,'o-')
    plt.plot(base,y_cs,'*-')
    plt.xlabel('base')
    plt.ylabel('Y')
    plt.grid(True)
    plt.show()

def twoD_view_static(x,y):
    fig = plt.figure()
    plt.xlabel('X')
    plt.ylabel('Y')
    plt.xlim([min(x)-1,max(x)+1])
    plt.ylim([min(y)-1,max(y)+1])
    plt.grid(True)
    plt.plot(x,y,'o-')
    plt.show()
   
def twoD_view_dynamic(x,y): 
    fig = plt.figure()
    plt.title('2D view dynamic plot')
    plt.grid(True)
    plt.axis([min(x)-1, max(x)+1, min(y)-1, max(y)+1]) 
    
    for i in range(0,len(x)):
        plt.plot(x[i],y[i],'ro')
        plt.pause(0.05)                  
        plt.show()  

def twoD_view_dynamic_cycles(x,y):
    seq_peak, seq_peak_id = min_minID(x)
    count = 0
    x_buff_temp = []
    y_buff_temp = []
    for i in range(0, len(x)):
        if(count < len(seq_peak_id)):
            if(i <= seq_peak_id[count]):
                x_buff_temp.append(x[i])  
                y_buff_temp.append(y[i]) 
            else:
                if(count>1):
                    fig = plt.figure()
                    ax = fig.add_subplot(111)
                    plt.grid(True)
                    plt.xlim([min(x),max(x)])
                    plt.ylim([min(y),max(y)])
                    for j in range(0,len(x_buff_temp)):
                        if(j == 0):
                            x_temp = x_buff_temp[j]  
                            y_temp = y_buff_temp[j] 
                        else: 
                            ax.quiver(x_temp,y_temp, (x_buff_temp[j]-x_temp),(y_buff_temp[j]-y_temp),angles='xy', scale_units='xy', scale =1)
                            plt.plot(x_buff_temp,y_buff_temp,'ro')
                        x_temp = x_buff_temp[j] 
                        y_temp = y_buff_temp[j]
                count = count + 1
                x_buff_temp = []
                y_buff_temp = []
                plt.pause(1)                  
                plt.show()
     
             
def twoD_view_dynamic_arrow(x,y): 
    x_temp = 0
    y_temp = 0
    fig = plt.figure()
    plt.title('2D view static plot with arrow')
    ax = fig.add_subplot(111)
    plt.grid(True)
    ax.axis([min(x)-1, max(x)+1, min(y)-1, max(y)+1]) 
    
    for i in range(0,len(x)):
        if(i == 0):
           x_temp = x[i]
           y_temp = y[i] 
        else:
            ax.quiver(x_temp,y_temp, (x[i]-x_temp),(y[i]-y_temp),angles='xy', scale_units='xy', scale =1)
            plt.plot(x_temp,y_temp,'ro')
            plt.plot(x[i],y[i],'ro')
            x_temp = x[i] 
            y_temp = y[i]
            plt.pause(0.1)                  
            plt.show()
 
def threeD_view_static(x,y,z):             
    fig = plt.figure()
    plt.grid(True)
    plt.title('3D view static plot')
    plt.ion()
    
    ax = fig.add_subplot(111, projection='3d')
    ax.set_xlabel('X Label')
    ax.set_ylabel('Y Label')
    ax.set_zlabel('Z Label')

    ax.set_xlim([min(x), max(x)])
    ax.set_ylim([min(y), max(y)])
    ax.set_zlim([min(z), max(z)])
    
    ax.scatter(x,y,z,color='r')
    plt.show()  
    

def threeD_view_dynamic(x,y,z): 
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')  
    for i in range(0, len(x)):
        ax.scatter(x[i], y[i], z[i],'ro')
        plt.show()
        plt.pause(0.05)
  
def threeD_view_dynamic_cycle(x,y,z): 
    count = 0
    xbuff = []
    ybuff = []
    zbuff = []
    seq_peak, seq_peak_id = min_minID(x)

    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')  
    for i in range(0, len(x)):
        if(count < len(seq_peak_id)):
            if(i <= seq_peak_id[count]):
                xbuff.append(x[i])  
                ybuff.append(y[i]) 
                zbuff.append(z[i])
            else:
                 ax.scatter(xbuff, ybuff, zbuff,'ro')
                 plt.plot(xbuff, ybuff, zbuff)
                 plt.show()
                 plt.pause(0.8)
                 xbuff = []
                 ybuff = []
                 zbuff = []
                 xbuff.append(x[i])  
                 ybuff.append(y[i])
                 zbuff.append(z[i])
                 count = count +1
   
class Arrow3D(FancyArrowPatch):
    def __init__(self, xs, ys, zs, *args, **kwargs):
        FancyArrowPatch.__init__(self, (0,0), (0,0), *args, **kwargs)
        self._verts3d = xs, ys, zs

    def draw(self, renderer):
        xs3d, ys3d, zs3d = self._verts3d
        xs, ys, zs = proj3d.proj_transform(xs3d, ys3d, zs3d, renderer.M)
        self.set_positions((xs[0],ys[0]),(xs[1],ys[1]))
        FancyArrowPatch.draw(self, renderer)
        
def threeD_view_dynamic_arrow(x,y,z):   
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')  

    ax.set_xlabel('X')
    ax.set_ylabel('Y')
    ax.set_zlabel('Z')

    ax.set_xlim([min(x), max(x)])
    ax.set_ylim([min(y), max(y)])
    ax.set_zlim([min(z), max(z)])

    tempx = 0
    tempy = 0
    tempz = 0
    for i in range(0, len(x)):
        ax.scatter(x[i],y[i],z[i],color='blue')
   
        if(tempx != 0):
            arw = Arrow3D([tempx,x[i]],[tempy,y[i]],[tempz,z[i]], arrowstyle="->", color="purple", lw = 1, mutation_scale=15)
            ax.add_artist(arw)
   
        tempx = x[i]
        tempy = y[i]
        tempz = z[i]
        plt.pause(0.001)                  
        plt.show()

def oneD_fft(y):
    N = len(y)
    # sample spacing
    T = 1.0 / 30
    base = np.linspace(0.0, N*T, N)

    #y = []
    #y = np.sin(50.0 * 2.0*np.pi*x) + 0.5*np.sin(80.0 * 2.0*np.pi*x)
    y = y - np.mean(y)
    plt.figure(1)
    plt.plot(base, y,'o-')
    yf = fft(y)
    xf = np.linspace(0.0, 1.0/(2.0*T), N/2)
    plt.figure(2)
    plt.plot(xf, 2.0/N * np.abs(yf[0:N/2]))
    plt.grid()
    plt.show()

if __name__ == "__main__":
    
#x,y,x_cs,y_cs = read_cs_abl_2d_data()
#x_base_graph(x)
    x,y,z = read_3d_data()

#x_base_graph(y)
#oneD_view_static_cs(x,y,x_cs,y_cs)
#twoD_view_static(x,y)
#twoD_view_dynamic_arrow(x,y)
    threeD_view_static(x,y,z)
#threeD_view_dynamic_arrow(x,y,z)


               

